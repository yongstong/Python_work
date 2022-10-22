
import openpyxl

from docx import Document


if __name__ == "__main__":

	workbook = openpyxl.load_workbook('question.xlsx')
	sheet = workbook.active
	doc = Document('2019-2020（B2）试卷.docx')
	for j in range(2,sheet.max_row,1):
		cell1 = sheet.cell(row=j, column=4)
		cella = sheet.cell(row=j, column=5)
		cellb = sheet.cell(row=j, column=6)
		cellc = sheet.cell(row=j, column=7)
		celld = sheet.cell(row=j, column=8)
		i = 0
		for paragraph in doc.paragraphs:
			for run in paragraph.runs:
				if '{题目}' in run.text:
					# 如果存在匹配得字符串，那么将当前得run替换成合并后得字符串
					run.text = run.text.replace('{题目}', str(cell1.value))
					i += 1

				elif '{选项A}' in run.text:
					# 如果存在匹配得字符串，那么将当前得run替换成合并后得字符串
					run.text = run.text.replace('{选项A}', str(cella.value))
					i += 1

				elif '{选项B}' in run.text:
					# 如果存在匹配得字符串，那么将当前得run替换成合并后得字符串
					run.text = run.text.replace('{选项B}', str(cellb.value))
					i += 1

				elif '{选项C}' in run.text:
					# 如果存在匹配得字符串，那么将当前得run替换成合并后得字符串
					run.text = run.text.replace('{选项C}', str(cellc.value))
					i += 1

				elif '{选项D}' in run.text:
					# 如果存在匹配得字符串，那么将当前得run替换成合并后得字符串
					run.text = run.text.replace('{选项D}', str(celld.value))
					i += 1
				if i == 5: break
			if i == 5: break
	doc.save('shijuan.docx')
