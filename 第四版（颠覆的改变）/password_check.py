import hashlib
import openpyxl


def get_md5(data):
    obj = hashlib.md5("e10adc3949ba59abbe56e057f20f883e".encode('utf-8'))
    obj.update(data.encode('utf-8'))
    result = obj.hexdigest()
    return result

def check_password(account,password):
    workbook = openpyxl.load_workbook('data/password.xlsx')
    sheet = workbook.active
    for i in range(sheet.max_row):
        cell_1=sheet.cell(row=i+1, column=1).value
        cell_2=sheet.cell(row=i+1, column=2).value
        if cell_1 == get_md5(account):
            if cell_2 == get_md5(account+password):
                print("1")
                return 1
            else:
                print("0")
                return 0
    print("-1")
    return -1
        
def register(account,password):
    workbook = openpyxl.load_workbook('data/password.xlsx')
    sheet = workbook.active
    max = sheet.max_row
    for i in range(sheet.max_row):
        if get_md5(account)==sheet.cell(row=i+1, column=1).value:
            return -1
    sheet.cell(row=max+1, column=1).value=get_md5(account)
    sheet.cell(row=max+1, column=2).value=get_md5(account+password)
    workbook.save('data/password.xlsx')
    print("over")
    return 1
if __name__ == "__main__":
    a = input("1登录,0注册")
    if a=='1':
        account = input("账号")
        password = input("密码")
        check_password(account,password)
    elif a=='0':
        account = input("账号")
        password = input("密码")
        register(account,password)